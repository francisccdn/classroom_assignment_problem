import os
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side


class borders:
    lr = Border(right=Side(border_style='thin', color='FF000000'), left=Side(border_style='thin', color='FF000000'))
    r = Border(right=Side(border_style='thin', color='FF000000'))
    b = Border(bottom=Side(border_style='thin', color='FF000000'))
    br = Border(bottom=Side(border_style='thin', color='FF000000'), right=Side(border_style='thin', color='FF000000'))
    blr = Border(bottom=Side(border_style='thin', color='FF000000'), right=Side(border_style='thin', color='FF000000'), left=Side(border_style='thin', color='FF000000'))
    tb = Border(top=Side(border_style='thin', color='FF000000'), bottom=Side(border_style='thin', color='FF000000'))
    tbl = Border(left=Side(border_style='thin', color='FF000000'), top=Side(border_style='thin', color='FF000000'), bottom=Side(border_style='thin', color='FF000000'))
    tbr = Border(right=Side(border_style='thin', color='FF000000'), top=Side(border_style='thin', color='FF000000'), bottom=Side(border_style='thin', color='FF000000'))
    tblr = Border(left=Side(border_style='thin', color='FF000000'), right=Side(border_style='thin', color='FF000000'), top=Side(border_style='thin', color='FF000000'), bottom=Side(border_style='thin', color='FF000000'))

class fonts:
    bold = Font(name="Calibri", bold=True)

def flatten(t):
    return [item for sublist in t for item in sublist]

def write_header(sheet, row):
    # Values
    sheet[f"A{row}"].value = "Instance"

    sheet[f"B{row}"].value = "Scenario"
    sheet[f"C{row}"].value = "Status"
    sheet[f"D{row}"].value = "Gap"
    sheet[f"E{row}"].value = "Solution value"    
    sheet[f"F{row}"].value = "Time reported at < 1%"
    sheet[f"G{row}"].value = "Time reported at < 0.1%"
    sheet[f"H{row}"].value = "Time: Solver"
    sheet[f"I{row}"].value = "Time: Model construction"
    sheet[f"J{row}"].value = "Time: Pre-processing"
    
    sheet[f"K{row}"].value = "Total kWh"
    sheet[f"L{row}"].value = "Total $"

    sheet[f"M{row}"].value = "Savings %"
    sheet[f"N{row}"].value = "Savings kWh"
    sheet[f"O{row}"].value = "Savings $"

    # Styles
    for cell in flatten(list(list(sheet[f"A{row}":f"O{row}"]))):
        cell.font = fonts.bold
        cell.border = borders.tb

    sheet[f"A{row}"].border = borders.tblr
    sheet[f"B{row}"].border = borders.tbl
    sheet[f"J{row}"].border = borders.tbr
    sheet[f"K{row}"].border = borders.tbl
    sheet[f"L{row}"].border = borders.tbr
    sheet[f"M{row}"].border = borders.tbl
    sheet[f"O{row}"].border = borders.tbr

def write_instance(sheet, start_row, instance):
    # Instance name
    sheet[f"A{start_row}"].value = instance["name"]
    sheet.merge_cells(range_string=f"A{start_row}:A{start_row+9}")
    sheet[f"A{start_row}"].font = fonts.bold
    sheet[f"A{start_row}"].border = borders.tblr
    sheet[f"A{start_row+9}"].border = borders.blr

    if instance["name"] == "20181":
        manual_col = "T"
    if instance["name"] == "20182":
        manual_col = "S"    
    elif instance["name"] == "20191":
        manual_col = "R"

    # Main content
    for i in range(1, 11):
        row = start_row + (i - 1)
        data = instance[i]

        sheet[f"B{row}"].value = data["scenario"]
        
        if data["status"] == 2:
            sheet[f"C{row}"].value = "Optimal"
        elif data["status"] == 1:
            sheet[f"C{row}"].value = "Feasible"
        else:
            sheet[f"C{row}"].value = data["status"]
        
        sheet[f"D{row}"].value = data["gap"]
        sheet[f"E{row}"].value = data["value"]

        if "time to 1%" in data:
            sheet[f"F{row}"].value = data["time to 1%"] if data["time to 1%"] != 9999999 else "NA"
        if "time to 0.1%" in data:
            sheet[f"G{row}"].value = data["time to 0.1%"] if data["time to 0.1%"] != 9999999 else "NA"

        sheet[f"H{row}"].value = data["time solver"]
        sheet[f"I{row}"].value = data["time model"]
        sheet[f"J{row}"].value = data["time pre processing"]

        sheet[f"K{row}"].value = f"=(((E{row}*50/60)/5)*100)"
        sheet[f"L{row}"].value = f"=(K{row}*0,82961)"

        sheet[f"M{row}"].value = f"=((E{row}-${manual_col}$6)/${manual_col}$6)*(-1)"
        sheet[f"N{row}"].value = f"=${manual_col}$7-K{row}"
        sheet[f"O{row}"].value = f"=${manual_col}$8-L{row}"
        
        # Borders
        if i == 10:
            for col in ["B", "C", "D", "E", "F", "G", "H", "I", "K", "M", "N"]:
                sheet[f"{col}{row}"].border = borders.b

        for col in ["J", "L", "O"]:
            if i != 10:
                sheet[f"{col}{row}"].border = borders.r
            else:
                sheet[f"{col}{row}"].border = borders.br
        
def write_manual_results(sheet, data):

    # Header
    sheet["Q4"].value = "MANUAL SOLUTIONS"
    sheet.merge_cells(range_string="Q4:T4")
    sheet["Q4"].font = fonts.bold
    sheet["Q4"].border = borders.tblr
    sheet["T4"].border = borders.tbr
    
    # Captions
    sheet["Q5"].value = "Instance"
    sheet["Q6"].value = "Cost"
    sheet["Q7"].value = "Total kWh"
    sheet["Q8"].value = "Total $"
    sheet["Q9"].value = "Num Assignments"
    sheet["Q10"].value = "Avg. Cost"
    sheet["Q11"].value = "Highest Cost - Clasrooms"
    sheet["Q12"].value = "Highest Cost - PC"

    for cell in flatten(list(list(sheet["Q6:Q11"]))):
        cell.font = fonts.bold
        cell.border = borders.lr
    for cell in [sheet["Q5"], sheet["Q12"]]:
        cell.font = fonts.bold
        cell.border = borders.blr

    # Content
    col = "R"
    for instance in data:
        sheet[f"{col}5"].value = instance
        sheet[f"{col}6"].value = data[instance]["setup_before"]
        sheet[f"{col}7"].value = f"=((({col}6*50/60)/5)*100)"
        sheet[f"{col}8"].value = f"=({col}7*0.82961)"
        sheet[f"{col}9"].value = data[instance]["n_assignments"]
        sheet[f"{col}10"].value = f"=({col}6/{col}9)"
        sheet[f"{col}11"].value = data[instance]["highest"]["room"]["setup_before"]
        sheet[f"{col}12"].value = data[instance]["highest"]["pc"]["setup_before"]

        for row in [5, 12]:
            sheet[f"{col}{row}"].border = borders.b
            if col == "T":
                sheet[f"{col}{row}"].border = borders.br
        
        col = chr(ord(col) + 1)
    
    for cell in flatten(list(list(sheet["T6:T11"]))):
        cell.border = borders.r
    
def get_results(path):
    data = {
        "20181": {
            "name": "20181"
        },
        "20182": {
            "name": "20182"
        }, 
        "20191": {
            "name": "20191"
        }
    }
    for file in os.listdir(path):
        filename = os.fsdecode(file)
        filepath = os.path.join(path, filename)

        with open(filepath, 'r') as fr:
            res = json.load(fr)
            data[res["instance"]][res["scenario"]] = res

    return data            


def main():
    # Create spreadsheet
    wb = Workbook()
    sheet = wb["Sheet"]
    # Write content headers
    write_header(sheet, 1)
    # Get results
    results_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.realpath(__file__)))) + '/results'
    data = get_results(results_dir)
    # Write instances' content
    row = 2
    for key in data:
        write_instance(sheet, row, data[key])
        row += 10
    # Write manual results
    manual_path = os.path.dirname(os.path.realpath(__file__)) + '/results_manual/manual_results.json'
    with open(manual_path, 'r') as fr:
        manual_data = json.load(fr)
        write_manual_results(sheet, manual_data)
    # Save spreasheet to file
    wb.save("cap_results.xlsx")

if __name__ == "__main__":
    main()