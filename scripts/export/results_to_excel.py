import os
import json
from openpyxl import Workbook

# Spreadsheet
wb = Workbook()

# Create a new sheet with given columns
def new_sheet(name: str):
    wb.create_sheet(name)
    sheet = wb[name]

    sheet["A1"].value = "Scenario"
    sheet["B1"].value = "Setup cost"
    sheet["C1"].value = "Setup before class"

    sheet["D1"].value = "Status"
    sheet["E1"].value = "Gap"
    sheet["F1"].value = "Solution value"
    
    sheet["G1"].value = "Time reported at < 1%"
    sheet["H1"].value = "Time reported at < 0.1%"
    
    sheet["I1"].value = "Time: Solver"
    sheet["J1"].value = "Time: Model construction"
    sheet["K1"].value = "Time: Pre-processing"
    
    wb.create_sheet("Heuristic - " + name)
    heur_sheet = wb["Heuristic - " + name]
    
    heur_sheet["A1"].value = "Scenario"
    heur_sheet["B1"].value = "Setup cost"
    heur_sheet["C1"].value = "Setup before class"

    heur_sheet["D1"].value = "Greedy Heuristic: Time"
    heur_sheet["E1"].value = "Greedy Heuristic: Num Unfeasibilities"
    heur_sheet["F1"].value = "Greedy Heuristic: Solution Value"
    heur_sheet["G1"].value = "Local Search: Time"
    heur_sheet["H1"].value = "Local Search: Solution Value"

# Looping through results 
indexes = [0] 
dir_path = os.path.dirname(os.path.dirname(os.path.dirname(os.path.realpath(__file__)))) + '/results'

for file in os.listdir(dir_path):
    filename = os.fsdecode(file)
    filepath = os.path.join(dir_path, filename)

    with open(filepath, 'r') as fr:
        data = json.load(fr)

        # Create a sheet for each instance
        if data['instance'] not in wb.sheetnames:
            new_sheet(data['instance'])
            indexes.append(2)
            indexes.append(2)
           
        # Index of indexes - Tells which row of instance sheet to write on
        j = wb.sheetnames.index(data['instance'])
        j_heur = wb.sheetnames.index("Heuristic - " + data['instance'])

        # Set working sheet
        sheet = wb[data['instance']]
        heur_sheet = wb["Heuristic - " + data['instance']]

        # Fill in row with file data
        if data["heuristic"] == False:
            sheet["A" + str(indexes[j])].value = data["scenario"]
            sheet["B" + str(indexes[j])].value = data["setup"]
            sheet["C" + str(indexes[j])].value = data["setup before class"]
            
            if data["status"] == 2:
                sheet["D" + str(indexes[j])].value = "Optimal"
            elif data["status"] == 1:
                sheet["D" + str(indexes[j])].value = "Feasible"
            else:
                sheet["D" + str(indexes[j])].value = data["status"]
            sheet["E" + str(indexes[j])].value = data["gap"]
            sheet["F" + str(indexes[j])].value = data["value"]
            
            sheet["G" + str(indexes[j])].value = data["time to 1%"]
            sheet["H" + str(indexes[j])].value = data["time to 0.1%"]
            
            sheet["I" + str(indexes[j])].value = data["time solver"]
            sheet["J" + str(indexes[j])].value = data["time model"]
            sheet["K" + str(indexes[j])].value = data["time pre processing"]

            indexes[j] += 1
        
        else:
            heur_sheet["A" + str(indexes[j_heur])].value = data["scenario"]
            heur_sheet["B" + str(indexes[j_heur])].value = data["setup"]
            heur_sheet["C" + str(indexes[j_heur])].value = data["setup before class"]
            
            heur_sheet["D" + str(indexes[j_heur])].value = data["heuristic - greedy - time"]
            heur_sheet["E" + str(indexes[j_heur])].value = data["heuristic - greedy - num unfeasibilities"]
            heur_sheet["F" + str(indexes[j_heur])].value = data["heuristic - greedy - value"]
            heur_sheet["G" + str(indexes[j_heur])].value = data["heuristic - local search - time"]
            heur_sheet["H" + str(indexes[j_heur])].value = data["heuristic - local search - value"]

            indexes[j_heur] += 1
        

# Get manual results
wb.create_sheet("Manual")
sheet = wb["Manual"]

sheet["A1"].value = "Instance"

sheet["A2"].value = "Cost: No setup"
sheet["A3"].value = "Cost: Setup before"
sheet["A4"].value = "Cost: Setup during"

sheet["A5"].value = "Num Assignments"

sheet["A6"].value = "Avg. Cost: No setup"
sheet["A7"].value = "Avg. Cost: Setup before"
sheet["A8"].value = "Avg. Cost: Setup during"

sheet["A9"].value = "Highest Cost - Clasrooms: No setup"
sheet["A10"].value = "Highest Cost - Clasrooms: Setup before"
sheet["A11"].value = "Highest Cost - Clasrooms: Setup during"
sheet["A12"].value = "Highest Cost - PC: No setup"
sheet["A13"].value = "Highest Cost - PC: Setup before"
sheet["A14"].value = "Highest Cost - PC: Setup during"

manual_path = os.path.dirname(os.path.realpath(__file__)) + '/results_manual/manual_results.json'
with open(manual_path, 'r') as fr:
    manual = json.load(fr)

    k = "B"
    for instance in manual:
        sheet[k + "1"].value = instance
        
        sheet[k + "2"].value = manual[instance]["no_setup"]
        sheet[k + "3"].value = manual[instance]["setup_before"]
        sheet[k + "4"].value = manual[instance]["setup_during"]

        sheet[k + "5"].value = manual[instance]["n_assignments"]
        
        sheet[k + "6"].value = f"=DIVIDE({k}2;{k}5)"
        sheet[k + "7"].value = f"=DIVIDE({k}3;{k}5)"
        sheet[k + "8"].value = f"=DIVIDE({k}4;{k}5)"
        
        sheet[k + "9"].value = manual[instance]["highest"]["room"]["no_setup"]
        sheet[k + "10"].value = manual[instance]["highest"]["room"]["setup_before"]
        sheet[k + "11"].value = manual[instance]["highest"]["room"]["setup_during"]
        sheet[k + "12"].value = manual[instance]["highest"]["pc"]["no_setup"]
        sheet[k + "13"].value = manual[instance]["highest"]["pc"]["setup_before"]
        sheet[k + "14"].value = manual[instance]["highest"]["pc"]["setup_during"]
        
        k = chr(ord(k) + 1)

# Remove default sheet
wb.remove(wb["Sheet"])
# Save spreasheet to file
wb.save("cap_results.xlsx")